from django.core.management.base import BaseCommand, CommandError
from market.models import Category, Product
from openpyxl import load_workbook

from prj.settings import DATA_DIR, BASE_DIR


CURRENT_DATA_DIR_PATH_FULL=fr'D:\GitHub\site_django_from_dima_herson\backend\init_data'

CURRENT_DATA_DIR = 'init_data'
CURRENT_DATA_DIR_PATH = fr'{BASE_DIR[:-3]}\{CURRENT_DATA_DIR}'                       
CURRENT_FILE_NAME = 'price.xlsx'

class Command(BaseCommand):


    def handle(self, *args, **options):
        print('Clearing DB')
        Category.objects.all().delete()
        Product.objects.all().delete()


        # current_path = f'{CURRENT_DATA_DIR_PATH_FULL}\\{CURRENT_FILE_NAME}'
        current_path = f'{CURRENT_DATA_DIR_PATH}\\{CURRENT_FILE_NAME}'

        print('start importing in excel %s' % current_path)

        # wb = load_workbook(DATA_DIR+'\price.xlsx')
        wb = load_workbook(current_path)
        current_sheet = wb.get_sheet_names()[0]
        
        sheet = (wb.get_sheet_by_name(current_sheet))
       
        current_cat = None
        for count in range(1, sheet.max_row+1):
            #----------------------------------------------------------
            # print('item', sheet.cell(row=count, column=2).value)
            # item = (sheet.cell(row=count, column=2).value)
            # if sheet.cell(row=count, column=1).value:
            #     print('id', sheet.cell(row=count, column=1).value)
            #     id = (sheet.cell(row=count, column=1).value)
            # else:
            #     print('category', sheet.cell(row=count, column=2).value)
            #     category = (sheet.cell(row=count, column=2).value)
            #--------------------------------------------------------
            item = (sheet.cell(row=count, column=2).value)
            id = (sheet.cell(row=count, column=1).value)
            if id ==None:
                print('Create a new category')
                cat = Category()
                cat.name = item
                cat.save()
                current_cat = cat
            else:    
                print('Create a new good')
                good = Product()
                good.name = item
                good.category = current_cat
                good.save()

                



